from __future__ import annotations

from dataclasses import asdict
import io
import math

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from engineering import retention, room_capacity, scanner_capacity
from models import (
    DevelopmentPhase,
    NetworkProfile,
    PlannerAssumptions,
    PlannerInputs,
    SharedNetworkAssumptions,
    SharedNetworkPhaseResult,
    SharedNetworkReport,
)
from optimization import conventional


NETWORK_ASSUMPTION_FIELDS = [
    "shared_backbone_cost",
    "guideway_segment_capex",
    "endpoint_capex",
    "vertical_transition_capex",
    "building_connection_capex",
    "new_room_construction_capex",
    "room_renovation_modification_capex",
    "room_connection_modification_capex",
    "production_expansion_capex_per_10pct",
    "cyclotron_purchase_capex",
    "cyclotron_installation_capex",
    "scanner_capex",
    "revenue_per_scan",
    "discount_rate_pct",
    "analysis_years",
    "operating_days_per_year",
    "operating_hours_per_day",
    "scanner_availability_pct",
    "scanner_cycle_min",
    "injection_cycle_min",
    "uptake_cycle_min",
    "conventional_transport_min",
    "mrt_transport_min",
]

NETWORK_ASSUMPTION_LABELS = {
    "shared_backbone_cost": "Shared MRT backbone cost",
    "guideway_segment_capex": "Guideway segment cost",
    "endpoint_capex": "Endpoint cost",
    "vertical_transition_capex": "Vertical transition cost",
    "building_connection_capex": "Building or floor connection cost",
    "new_room_construction_capex": "New room construction cost",
    "room_renovation_modification_capex": "Existing room renovation or modification cost",
    "room_connection_modification_capex": "MRT room connection or modification cost",
    "production_expansion_capex_per_10pct": "Production expansion cost per 10% block",
    "cyclotron_purchase_capex": "Cyclotron purchase cost",
    "cyclotron_installation_capex": "Cyclotron installation cost",
    "scanner_capex": "Scanner cost",
    "revenue_per_scan": "Revenue per scan",
    "discount_rate_pct": "Discount rate",
    "analysis_years": "Analysis period",
    "operating_days_per_year": "Operating days per year",
    "operating_hours_per_day": "Operating hours per day",
    "scanner_availability_pct": "Scanner availability",
    "scanner_cycle_min": "Scanner cycle time",
    "injection_cycle_min": "Injection room cycle time",
    "uptake_cycle_min": "Uptake room cycle time",
    "conventional_transport_min": "Conventional transport time",
    "mrt_transport_min": "MRT transport time",
}


def default_network_profile() -> NetworkProfile:
    return NetworkProfile(
        study_name="Campus Shared Network Study",
        facility_name="Main Medical Campus",
        baseline_current_patients_per_day=100.0,
        baseline_max_expected_demand_per_day=100.0,
        baseline_scanners=3,
        baseline_injection_rooms=6,
        baseline_uptake_rooms=6,
        baseline_usable_doses_per_day=120.0,
        has_existing_cyclotron=True,
        existing_backbone_installed=False,
        initial_guideway_segments=0,
        initial_endpoints=0,
        initial_connected_rooms=0,
        initial_vertical_transitions=0,
        initial_building_connections=0,
    )


def default_development_phases() -> list[DevelopmentPhase]:
    return [
        DevelopmentPhase(
            phase_name="Phase 1",
            year=0,
            service_group="PET Expansion",
            representative_radionuclide="F-18",
            incremental_target_patients_per_day=80.0,
            maximum_expected_demand_per_day=80.0,
            existing_rooms_to_connect=4,
            new_rooms_to_construct=2,
            cumulative_guideway_segments_required=3,
            cumulative_endpoints_required=8,
            cumulative_vertical_transitions_required=1,
            cumulative_building_connections_required=0,
        ),
        DevelopmentPhase(
            phase_name="Phase 2",
            year=2,
            service_group="Theranostics",
            representative_radionuclide="Ga-68",
            incremental_target_patients_per_day=50.0,
            maximum_expected_demand_per_day=50.0,
            existing_rooms_to_connect=3,
            new_rooms_to_construct=2,
            cumulative_guideway_segments_required=5,
            cumulative_endpoints_required=12,
            cumulative_vertical_transitions_required=1,
            cumulative_building_connections_required=1,
        ),
        DevelopmentPhase(
            phase_name="Phase 3",
            year=4,
            service_group="SPECT / NM",
            representative_radionuclide="Tc-99m",
            incremental_target_patients_per_day=40.0,
            maximum_expected_demand_per_day=40.0,
            existing_rooms_to_connect=2,
            new_rooms_to_construct=1,
            cumulative_guideway_segments_required=6,
            cumulative_endpoints_required=15,
            cumulative_vertical_transitions_required=2,
            cumulative_building_connections_required=1,
        ),
    ]


def default_network_assumption_values() -> dict[str, float]:
    defaults = SharedNetworkAssumptions()
    values = asdict(defaults)
    values["analysis_years"] = float(values["analysis_years"])
    values["operating_days_per_year"] = float(values["operating_days_per_year"])
    return values


def assumptions_from_values(values: dict[str, float]) -> SharedNetworkAssumptions:
    return SharedNetworkAssumptions(
        shared_backbone_cost=float(values["shared_backbone_cost"]),
        guideway_segment_capex=float(values["guideway_segment_capex"]),
        endpoint_capex=float(values["endpoint_capex"]),
        vertical_transition_capex=float(values["vertical_transition_capex"]),
        building_connection_capex=float(values["building_connection_capex"]),
        new_room_construction_capex=float(values["new_room_construction_capex"]),
        room_renovation_modification_capex=float(values["room_renovation_modification_capex"]),
        room_connection_modification_capex=float(values["room_connection_modification_capex"]),
        production_expansion_capex_per_10pct=float(values["production_expansion_capex_per_10pct"]),
        cyclotron_purchase_capex=float(values["cyclotron_purchase_capex"]),
        cyclotron_installation_capex=float(values["cyclotron_installation_capex"]),
        scanner_capex=float(values["scanner_capex"]),
        revenue_per_scan=float(values["revenue_per_scan"]),
        discount_rate_pct=float(values["discount_rate_pct"]),
        analysis_years=int(values["analysis_years"]),
        operating_days_per_year=int(values["operating_days_per_year"]),
        operating_hours_per_day=float(values["operating_hours_per_day"]),
        scanner_availability_pct=float(values["scanner_availability_pct"]),
        scanner_cycle_min=float(values["scanner_cycle_min"]),
        injection_cycle_min=float(values["injection_cycle_min"]),
        uptake_cycle_min=float(values["uptake_cycle_min"]),
        conventional_transport_min=float(values["conventional_transport_min"]),
        mrt_transport_min=float(values["mrt_transport_min"]),
    )


def changed_network_assumption_labels(values: dict[str, float]) -> list[str]:
    defaults = default_network_assumption_values()
    changed: list[str] = []
    for key in NETWORK_ASSUMPTION_FIELDS:
        if abs(float(values[key]) - float(defaults[key])) > 1e-9:
            changed.append(NETWORK_ASSUMPTION_LABELS[key])
    return changed


def validate_network_assumptions(values: dict[str, float]) -> list[str]:
    issues: list[str] = []
    for key in NETWORK_ASSUMPTION_FIELDS:
        if values[key] < 0:
            issues.append(f"{NETWORK_ASSUMPTION_LABELS[key]} cannot be negative.")

    if values["discount_rate_pct"] > 100:
        issues.append("Discount rate must be between 0% and 100%.")
    if values["analysis_years"] < 1:
        issues.append("Analysis period must be at least 1 year.")
    if values["operating_days_per_year"] < 1 or values["operating_days_per_year"] > 366:
        issues.append("Operating days per year must be between 1 and 366.")
    if values["scanner_availability_pct"] <= 0 or values["scanner_availability_pct"] > 100:
        issues.append("Scanner availability must be greater than 0% and at most 100%.")
    return issues


def validate_shared_inputs(profile: NetworkProfile, phases: list[DevelopmentPhase]) -> list[str]:
    issues: list[str] = []
    if not profile.study_name.strip():
        issues.append("Study name is required.")
    if not profile.facility_name.strip():
        issues.append("Facility name is required.")
    if profile.baseline_current_patients_per_day <= 0:
        issues.append("Baseline current patients/day must be positive.")
    if profile.baseline_max_expected_demand_per_day <= 0:
        issues.append("Baseline maximum expected demand/day must be positive.")
    if profile.baseline_scanners < 0 or profile.baseline_injection_rooms < 0 or profile.baseline_uptake_rooms < 0:
        issues.append("Baseline scanner and room counts cannot be negative.")
    if profile.baseline_usable_doses_per_day <= 0:
        issues.append("Baseline usable doses/day must be positive.")

    if len(phases) < 3:
        issues.append("Shared-network study requires at least 3 phases.")

    seen_names: set[str] = set()
    last_year = -1
    for phase in sorted(phases, key=lambda p: (p.year, p.phase_name)):
        if phase.phase_name in seen_names:
            issues.append("Phase names must be unique.")
        seen_names.add(phase.phase_name)
        if phase.year < 0:
            issues.append("Phase year must be zero or greater.")
        if phase.year < last_year:
            issues.append("Phase years must be non-decreasing.")
        last_year = phase.year
        if phase.incremental_target_patients_per_day < 0:
            issues.append("Incremental target patients/day must be zero or greater.")
        if phase.maximum_expected_demand_per_day < 0:
            issues.append("Maximum expected demand/day must be zero or greater.")
        if phase.existing_rooms_to_connect < 0 or phase.new_rooms_to_construct < 0:
            issues.append("Room quantities cannot be negative.")
        if phase.existing_rooms_to_renovate < 0 or phase.conventional_existing_rooms_to_renovate < 0:
            issues.append("Existing room renovation quantities cannot be negative.")
        if phase.mrt_new_rooms_requiring_connection_modification < 0:
            issues.append("MRT new-room connection-modification quantity cannot be negative.")
        if phase.mrt_new_rooms_requiring_connection_modification > phase.new_rooms_to_construct:
            issues.append("MRT new rooms requiring connection modification cannot exceed explicit new rooms to construct.")
        if phase.cumulative_guideway_segments_required < 0:
            issues.append("Cumulative guideway requirement cannot be negative.")
        if phase.cumulative_endpoints_required < 0:
            issues.append("Cumulative endpoint requirement cannot be negative.")
        if phase.cumulative_vertical_transitions_required < 0:
            issues.append("Cumulative vertical-transition requirement cannot be negative.")
        if phase.cumulative_building_connections_required < 0:
            issues.append("Cumulative building-connection requirement cannot be negative.")

    return issues


def _planner_assumptions(shared: SharedNetworkAssumptions) -> PlannerAssumptions:
    return PlannerAssumptions(
        mrt_infrastructure_capex=shared.shared_backbone_cost,
        cyclotron_purchase_capex=shared.cyclotron_purchase_capex,
        cyclotron_installation_capex=shared.cyclotron_installation_capex,
        additional_room_capex=shared.new_room_construction_capex,
        production_expansion_capex_per_10pct=shared.production_expansion_capex_per_10pct,
        discount_rate_pct=shared.discount_rate_pct,
        analysis_years=shared.analysis_years,
        operating_days_per_year=shared.operating_days_per_year,
        scanner_availability_pct=shared.scanner_availability_pct,
        scanner_cycle_min=shared.scanner_cycle_min,
        injection_cycle_min=shared.injection_cycle_min,
        uptake_cycle_min=shared.uptake_cycle_min,
        operating_hours_per_day=shared.operating_hours_per_day,
        mrt_transport_default_min=shared.mrt_transport_min,
        scanner_capex=shared.scanner_capex,
        endpoint_capex=shared.endpoint_capex,
        guideway_segment_capex=shared.guideway_segment_capex,
        revenue_per_scan=shared.revenue_per_scan,
    )


def _lookup_half_life(phase: DevelopmentPhase, half_life_lookup: dict[str, float]) -> float:
    if phase.representative_radionuclide in half_life_lookup:
        return float(half_life_lookup[phase.representative_radionuclide])
    return float(half_life_lookup.get("F-18", 109.8))


def _build_inputs(
    state: dict[str, float | int | bool],
    phase: DevelopmentPhase,
    profile: NetworkProfile,
    target_patients: float,
    max_demand: float,
    transport_min: float,
) -> PlannerInputs:
    return PlannerInputs(
        project_name=profile.study_name,
        current_patients_per_day=float(state["current_patients"]),
        target_patients_per_day=target_patients,
        maximum_expected_demand_per_day=max_demand,
        current_scanners=int(state["scanners"]),
        current_injection_rooms=int(state["injection_rooms"]),
        current_uptake_rooms=int(state["uptake_rooms"]),
        has_existing_cyclotron=bool(state["has_cyclotron"]),
        current_usable_doses_per_day=float(state["usable_doses"]),
        current_average_transport_min=transport_min,
        mrt_transport_min=None,
        existing_mrt_connectable_rooms=0,
        representative_radionuclide=phase.representative_radionuclide,
        representative_half_life_min=None,
    )


def _room_unit_capacity(assumptions: SharedNetworkAssumptions) -> float:
    return room_capacity(1, assumptions.operating_hours_per_day, assumptions.uptake_cycle_min)


def _network_utilization(supported_patients: float, connected_rooms: int, assumptions: SharedNetworkAssumptions) -> float:
    if connected_rooms <= 0:
        return 0.0
    cap = connected_rooms * _room_unit_capacity(assumptions)
    if cap <= 0:
        return 0.0
    return min(100.0, max(0.0, supported_patients / cap * 100.0))


def _append_ledger_row(
    ledger: list[dict[str, object]],
    phase: DevelopmentPhase,
    pathway: str,
    component: str,
    opening: float,
    required_total: float,
    incremental: float,
    unit_cost: float,
    cumulative_subtotal: float,
    formula_basis: str,
    assumption_source: str,
    extra: dict[str, object] | None = None,
) -> None:
    payload = {
        "phase": phase.phase_name,
        "year": phase.year,
        "pathway": pathway,
        "service_group": phase.service_group,
        "component": component,
        "opening_quantity_or_state": opening,
        "required_total_quantity_or_state": required_total,
        "incremental_quantity": incremental,
        "unit_cost": unit_cost,
        "phase_subtotal": incremental * unit_cost,
        "cumulative_subtotal": cumulative_subtotal,
        "formula_basis": formula_basis,
        "assumption_source": assumption_source,
    }
    if extra:
        payload.update(extra)
    ledger.append(payload)


def run_shared_network_study(
    profile: NetworkProfile,
    phases: list[DevelopmentPhase],
    assumptions: SharedNetworkAssumptions,
    half_life_lookup: dict[str, float],
) -> SharedNetworkReport:
    planner_assumptions = _planner_assumptions(assumptions)
    phases_sorted = sorted(phases, key=lambda p: (p.year, p.phase_name))

    conventional_state: dict[str, float | int | bool] = {
        "current_patients": profile.baseline_current_patients_per_day,
        "scanners": profile.baseline_scanners,
        "injection_rooms": profile.baseline_injection_rooms,
        "uptake_rooms": profile.baseline_uptake_rooms,
        "usable_doses": profile.baseline_usable_doses_per_day,
        "has_cyclotron": profile.has_existing_cyclotron,
    }
    mrt_state: dict[str, float | int | bool] = {
        "current_patients": profile.baseline_current_patients_per_day,
        "scanners": profile.baseline_scanners,
        "injection_rooms": profile.baseline_injection_rooms,
        "uptake_rooms": profile.baseline_uptake_rooms,
        "usable_doses": profile.baseline_usable_doses_per_day,
        "has_cyclotron": profile.has_existing_cyclotron,
    }

    backbone_installed = profile.existing_backbone_installed
    cumulative_guideway_segments = profile.initial_guideway_segments
    cumulative_endpoints = profile.initial_endpoints
    cumulative_connected_rooms = profile.initial_connected_rooms
    cumulative_vertical_transitions = profile.initial_vertical_transitions
    cumulative_building_connections = profile.initial_building_connections

    cumulative_target_patients = profile.baseline_current_patients_per_day
    cumulative_max_demand = profile.baseline_max_expected_demand_per_day

    cumulative_conventional_capex = 0.0
    cumulative_mrt_capex = 0.0
    cumulative_conventional_revenue = 0.0
    cumulative_mrt_revenue = 0.0

    prior_conventional_cumulative_revenue = 0.0
    prior_mrt_cumulative_revenue = 0.0

    phase_results: list[SharedNetworkPhaseResult] = []
    phase_ledger: list[dict[str, object]] = []
    network_state_rows: list[dict[str, object]] = []

    connected_services: set[str] = set()

    for phase in phases_sorted:
        connected_services.add(phase.service_group)

        cumulative_target_patients += phase.incremental_target_patients_per_day
        cumulative_max_demand += phase.maximum_expected_demand_per_day
        half_life_min = _lookup_half_life(phase, half_life_lookup)

        conv_inputs = _build_inputs(
            conventional_state,
            phase,
            profile,
            target_patients=cumulative_target_patients,
            max_demand=cumulative_max_demand,
            transport_min=assumptions.conventional_transport_min,
        )
        conv_plan = conventional(conv_inputs, planner_assumptions, half_life_min)
        conv_prod_blocks = math.ceil(conv_plan.required_production_increase_pct / 10.0) if conv_plan.required_production_increase_pct > 0 else 0

        conv_open_scanners = int(conventional_state["scanners"])
        conv_open_injection = int(conventional_state["injection_rooms"])
        conv_open_uptake = int(conventional_state["uptake_rooms"])
        conv_open_prod_blocks = 0.0

        conv_required_room_increment = conv_plan.additional_injection_rooms + conv_plan.additional_uptake_rooms
        conv_explicit_renovations = max(0, phase.conventional_existing_rooms_to_renovate)
        conv_auto_new_rooms = max(0, conv_required_room_increment - conv_explicit_renovations)
        conv_new_rooms_constructed = conv_auto_new_rooms
        conv_rooms_renovated = min(conv_required_room_increment, conv_explicit_renovations)

        forced_conv_cyclotron = phase.conventional_new_cyclotron_required and not bool(conventional_state["has_cyclotron"])
        conv_room_cost_in_plan = conv_required_room_increment * assumptions.new_room_construction_capex
        conv_room_cost_classified = (
            conv_new_rooms_constructed * assumptions.new_room_construction_capex
            + conv_rooms_renovated * assumptions.room_renovation_modification_capex
        )
        conv_phase_capex = conv_plan.capex - conv_room_cost_in_plan + conv_room_cost_classified
        if forced_conv_cyclotron and not conv_plan.cyclotron_required:
            conv_phase_capex += assumptions.cyclotron_purchase_capex + assumptions.cyclotron_installation_capex

        conventional_state["scanners"] = conv_open_scanners + conv_plan.additional_scanners
        conventional_state["injection_rooms"] = conv_open_injection + conv_plan.additional_injection_rooms
        conventional_state["uptake_rooms"] = conv_open_uptake + conv_plan.additional_uptake_rooms
        conventional_state["usable_doses"] = float(conventional_state["usable_doses"]) * (1.0 + conv_prod_blocks * 0.1)
        conventional_state["current_patients"] = cumulative_target_patients
        if conv_plan.cyclotron_required or forced_conv_cyclotron:
            conventional_state["has_cyclotron"] = True

        conv_throughput = min(conv_plan.achieved_capacity_per_day, cumulative_max_demand)
        conv_cumulative_revenue_now = conv_throughput * assumptions.revenue_per_scan * assumptions.operating_days_per_year
        conv_phase_revenue = conv_cumulative_revenue_now - prior_conventional_cumulative_revenue
        prior_conventional_cumulative_revenue = conv_cumulative_revenue_now

        cumulative_conventional_capex += conv_phase_capex
        cumulative_conventional_revenue += conv_phase_revenue

        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Scanners",
            float(conv_open_scanners),
            float(conventional_state["scanners"]),
            float(conv_plan.additional_scanners),
            assumptions.scanner_capex,
            cumulative_conventional_capex,
            "required_total_after_phase - opening_scanners",
            "SharedNetworkAssumptions.scanner_capex",
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Injection rooms (capacity result)",
            float(conv_open_injection),
            float(conventional_state["injection_rooms"]),
            float(conv_plan.additional_injection_rooms),
            0.0,
            cumulative_conventional_capex,
            "required_total_after_phase - opening_injection_rooms",
            "Capacity-only ledger row; construction cost tracked in classified room-cost rows",
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Uptake rooms (capacity result)",
            float(conv_open_uptake),
            float(conventional_state["uptake_rooms"]),
            float(conv_plan.additional_uptake_rooms),
            0.0,
            cumulative_conventional_capex,
            "required_total_after_phase - opening_uptake_rooms",
            "Capacity-only ledger row; construction cost tracked in classified room-cost rows",
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "New clinical room construction",
            0.0,
            float(conv_new_rooms_constructed),
            float(conv_new_rooms_constructed),
            assumptions.new_room_construction_capex,
            cumulative_conventional_capex,
            "max(required_room_increment - explicitly_renovated_rooms, 0)",
            "SharedNetworkAssumptions.new_room_construction_capex",
            extra={
                "cost_basis_category": "new_room_construction",
                "phase_first_installed": phase.phase_name,
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Existing room renovation or modification",
            0.0,
            float(conv_rooms_renovated),
            float(conv_rooms_renovated),
            assumptions.room_renovation_modification_capex,
            cumulative_conventional_capex,
            "min(required_room_increment, explicitly_renovated_rooms)",
            "SharedNetworkAssumptions.room_renovation_modification_capex",
            extra={
                "cost_basis_category": "room_renovation_modification",
                "phase_first_installed": phase.phase_name,
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Production expansion blocks (10%)",
            conv_open_prod_blocks,
            conv_open_prod_blocks + float(conv_prod_blocks),
            float(conv_prod_blocks),
            assumptions.production_expansion_capex_per_10pct,
            cumulative_conventional_capex,
            "ceil(required_production_increase_pct / 10)",
            "SharedNetworkAssumptions.production_expansion_capex_per_10pct",
        )
        conv_cyclotron_qty = 1.0 if (conv_plan.cyclotron_required or forced_conv_cyclotron) else 0.0
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Cyclotron purchase",
            0.0,
            conv_cyclotron_qty,
            conv_cyclotron_qty,
            assumptions.cyclotron_purchase_capex,
            cumulative_conventional_capex,
            "charged only if cyclotron required and not already installed",
            "SharedNetworkAssumptions.cyclotron_purchase_capex",
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "Conventional",
            "Cyclotron installation",
            0.0,
            conv_cyclotron_qty,
            conv_cyclotron_qty,
            assumptions.cyclotron_installation_capex,
            cumulative_conventional_capex,
            "charged only if cyclotron required and not already installed",
            "SharedNetworkAssumptions.cyclotron_installation_capex",
        )

        mrt_inputs = _build_inputs(
            mrt_state,
            phase,
            profile,
            target_patients=cumulative_target_patients,
            max_demand=cumulative_max_demand,
            transport_min=assumptions.mrt_transport_min,
        )
        mrt_plan_core = conventional(mrt_inputs, planner_assumptions, half_life_min)

        mrt_open_scanners = int(mrt_state["scanners"])
        mrt_open_injection = int(mrt_state["injection_rooms"])
        mrt_open_uptake = int(mrt_state["uptake_rooms"])

        required_room_increment = mrt_plan_core.additional_injection_rooms + mrt_plan_core.additional_uptake_rooms

        existing_rooms_connected = max(0, phase.existing_rooms_to_connect)
        existing_rooms_renovated = max(0, phase.existing_rooms_to_renovate)

        explicit_new_rooms = max(0, phase.new_rooms_to_construct)
        covered_by_explicit = existing_rooms_connected + existing_rooms_renovated + explicit_new_rooms
        auto_new_rooms = max(0, required_room_increment - covered_by_explicit)
        new_rooms_constructed = explicit_new_rooms + auto_new_rooms

        if phase.new_rooms_require_connection_modification:
            explicit_new_room_modifications = max(0, phase.mrt_new_rooms_requiring_connection_modification)
            new_room_connection_modifications = min(new_rooms_constructed, explicit_new_room_modifications)
        else:
            new_room_connection_modifications = 0

        room_connection_qty = existing_rooms_connected + new_room_connection_modifications

        scanner_increment = max(mrt_plan_core.additional_scanners, phase.additional_scanners_manual)

        required_guideway_total = max(cumulative_guideway_segments, phase.cumulative_guideway_segments_required)
        required_endpoints_total = max(cumulative_endpoints, phase.cumulative_endpoints_required)
        required_vertical_total = max(cumulative_vertical_transitions, phase.cumulative_vertical_transitions_required)
        required_building_total = max(cumulative_building_connections, phase.cumulative_building_connections_required)

        new_guideway = required_guideway_total - cumulative_guideway_segments
        new_endpoints = required_endpoints_total - cumulative_endpoints
        new_vertical = required_vertical_total - cumulative_vertical_transitions
        new_building = required_building_total - cumulative_building_connections

        mrt_prod_blocks = math.ceil(mrt_plan_core.required_production_increase_pct / 10.0) if mrt_plan_core.required_production_increase_pct > 0 else 0

        backbone_charged_this_phase = not backbone_installed
        mrt_cyclotron_charge = (mrt_plan_core.cyclotron_required or phase.mrt_new_cyclotron_required) and not bool(mrt_state["has_cyclotron"])

        mrt_phase_capex = 0.0
        if backbone_charged_this_phase:
            mrt_phase_capex += assumptions.shared_backbone_cost

        mrt_phase_capex += new_guideway * assumptions.guideway_segment_capex
        mrt_phase_capex += new_endpoints * assumptions.endpoint_capex
        mrt_phase_capex += new_vertical * assumptions.vertical_transition_capex
        mrt_phase_capex += new_building * assumptions.building_connection_capex
        mrt_phase_capex += new_rooms_constructed * assumptions.new_room_construction_capex
        mrt_phase_capex += room_connection_qty * assumptions.room_connection_modification_capex
        mrt_phase_capex += scanner_increment * assumptions.scanner_capex
        mrt_phase_capex += mrt_prod_blocks * assumptions.production_expansion_capex_per_10pct
        if mrt_cyclotron_charge:
            mrt_phase_capex += assumptions.cyclotron_purchase_capex + assumptions.cyclotron_installation_capex

        mrt_state["scanners"] = mrt_open_scanners + scanner_increment
        mrt_state["injection_rooms"] = mrt_open_injection + mrt_plan_core.additional_injection_rooms
        mrt_state["uptake_rooms"] = mrt_open_uptake + mrt_plan_core.additional_uptake_rooms
        mrt_state["usable_doses"] = float(mrt_state["usable_doses"]) * (1.0 + mrt_prod_blocks * 0.1)
        mrt_state["current_patients"] = cumulative_target_patients
        if mrt_cyclotron_charge:
            mrt_state["has_cyclotron"] = True

        cumulative_guideway_segments = required_guideway_total
        cumulative_endpoints = required_endpoints_total
        cumulative_vertical_transitions = required_vertical_total
        cumulative_building_connections = required_building_total
        cumulative_connected_rooms += existing_rooms_connected + new_rooms_constructed
        backbone_installed = True

        mrt_throughput = min(mrt_plan_core.achieved_capacity_per_day, cumulative_max_demand)
        mrt_cumulative_revenue_now = mrt_throughput * assumptions.revenue_per_scan * assumptions.operating_days_per_year
        mrt_phase_revenue = mrt_cumulative_revenue_now - prior_mrt_cumulative_revenue
        prior_mrt_cumulative_revenue = mrt_cumulative_revenue_now

        cumulative_mrt_capex += mrt_phase_capex
        cumulative_mrt_revenue += mrt_phase_revenue

        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "MRT base infrastructure",
            0.0 if not backbone_charged_this_phase else 0.0,
            1.0 if backbone_charged_this_phase else 1.0,
            1.0 if backbone_charged_this_phase else 0.0,
            assumptions.shared_backbone_cost,
            cumulative_mrt_capex,
            "charged once when shared backbone is first installed",
            "SharedNetworkAssumptions.shared_backbone_cost",
            extra={
                "mrt_backbone_charged_this_phase": backbone_charged_this_phase,
                "cost_basis_category": "shared_backbone_one_time",
                "phase_first_installed": phase.phase_name if backbone_charged_this_phase else "prior_phase",
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Guideway segments",
            float(required_guideway_total - new_guideway),
            float(required_guideway_total),
            float(new_guideway),
            assumptions.guideway_segment_capex,
            cumulative_mrt_capex,
            "required_cumulative_guideway - opening_guideway",
            "SharedNetworkAssumptions.guideway_segment_capex",
            extra={
                "cost_basis_category": "network_infrastructure_incremental",
                "phase_first_installed": phase.phase_name if new_guideway > 0 else "prior_phase",
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Endpoints",
            float(required_endpoints_total - new_endpoints),
            float(required_endpoints_total),
            float(new_endpoints),
            assumptions.endpoint_capex,
            cumulative_mrt_capex,
            "required_cumulative_endpoints - opening_endpoints",
            "SharedNetworkAssumptions.endpoint_capex",
            extra={
                "cost_basis_category": "network_infrastructure_incremental",
                "phase_first_installed": phase.phase_name if new_endpoints > 0 else "prior_phase",
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Vertical transitions",
            float(required_vertical_total - new_vertical),
            float(required_vertical_total),
            float(new_vertical),
            assumptions.vertical_transition_capex,
            cumulative_mrt_capex,
            "required_cumulative_vertical_transitions - opening_vertical_transitions",
            "SharedNetworkAssumptions.vertical_transition_capex",
            extra={
                "cost_basis_category": "network_infrastructure_incremental",
                "phase_first_installed": phase.phase_name if new_vertical > 0 else "prior_phase",
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Building or floor connections",
            float(required_building_total - new_building),
            float(required_building_total),
            float(new_building),
            assumptions.building_connection_capex,
            cumulative_mrt_capex,
            "required_cumulative_building_connections - opening_building_connections",
            "SharedNetworkAssumptions.building_connection_capex",
            extra={
                "cost_basis_category": "network_infrastructure_incremental",
                "phase_first_installed": phase.phase_name if new_building > 0 else "prior_phase",
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "New room construction",
            0.0,
            float(new_rooms_constructed),
            float(new_rooms_constructed),
            assumptions.new_room_construction_capex,
            cumulative_mrt_capex,
            "new rooms this phase + auto-added rooms required for physical feasibility",
            "SharedNetworkAssumptions.new_room_construction_capex",
            extra={
                "cost_basis_category": "new_room_construction",
                "phase_first_installed": phase.phase_name,
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Existing room renovation or modification",
            0.0,
            float(existing_rooms_renovated),
            float(existing_rooms_renovated),
            assumptions.room_renovation_modification_capex,
            cumulative_mrt_capex,
            "phase-declared existing rooms requiring conversion or renovation",
            "SharedNetworkAssumptions.room_renovation_modification_capex",
            extra={
                "cost_basis_category": "room_renovation_modification",
                "phase_first_installed": phase.phase_name,
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Room connection or modification",
            0.0,
            float(room_connection_qty),
            float(room_connection_qty),
            assumptions.room_connection_modification_capex,
            cumulative_mrt_capex,
            "existing rooms connected + optionally new rooms requiring MRT modification",
            "SharedNetworkAssumptions.room_connection_modification_capex",
            extra={
                "cost_basis_category": "mrt_connection_or_retrofit",
                "phase_first_installed": phase.phase_name,
                "existing_rooms_connected": existing_rooms_connected,
                "new_rooms_connection_modified": new_room_connection_modifications,
            },
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Scanners",
            float(mrt_open_scanners),
            float(mrt_state["scanners"]),
            float(scanner_increment),
            assumptions.scanner_capex,
            cumulative_mrt_capex,
            "max(required_scanner_increment, manual_scanner_increment)",
            "SharedNetworkAssumptions.scanner_capex",
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Production expansion blocks (10%)",
            0.0,
            float(mrt_prod_blocks),
            float(mrt_prod_blocks),
            assumptions.production_expansion_capex_per_10pct,
            cumulative_mrt_capex,
            "ceil(required_production_increase_pct / 10)",
            "SharedNetworkAssumptions.production_expansion_capex_per_10pct",
        )
        mrt_cyclotron_qty = 1.0 if mrt_cyclotron_charge else 0.0
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Cyclotron purchase",
            0.0,
            mrt_cyclotron_qty,
            mrt_cyclotron_qty,
            assumptions.cyclotron_purchase_capex,
            cumulative_mrt_capex,
            "charged only if cyclotron required and not already installed",
            "SharedNetworkAssumptions.cyclotron_purchase_capex",
        )
        _append_ledger_row(
            phase_ledger,
            phase,
            "MRT",
            "Cyclotron installation",
            0.0,
            mrt_cyclotron_qty,
            mrt_cyclotron_qty,
            assumptions.cyclotron_installation_capex,
            cumulative_mrt_capex,
            "charged only if cyclotron required and not already installed",
            "SharedNetworkAssumptions.cyclotron_installation_capex",
        )

        phase_capex_difference = mrt_phase_capex - conv_phase_capex
        cumulative_capex_difference = cumulative_mrt_capex - cumulative_conventional_capex
        cumulative_economic_difference = (
            cumulative_mrt_revenue - cumulative_conventional_revenue
        ) - (
            cumulative_capex_difference
        )

        network_utilization_pct = _network_utilization(
            mrt_plan_core.achieved_capacity_per_day,
            cumulative_connected_rooms,
            assumptions,
        )

        phase_results.append(
            SharedNetworkPhaseResult(
                phase_name=phase.phase_name,
                year=phase.year,
                service_group=phase.service_group,
                conventional_incremental_capex=conv_phase_capex,
                mrt_incremental_capex=mrt_phase_capex,
                phase_capex_difference=phase_capex_difference,
                conventional_cumulative_capex=cumulative_conventional_capex,
                mrt_cumulative_capex=cumulative_mrt_capex,
                cumulative_capex_difference=cumulative_capex_difference,
                conventional_phase_revenue=conv_phase_revenue,
                mrt_phase_revenue=mrt_phase_revenue,
                conventional_cumulative_revenue=cumulative_conventional_revenue,
                mrt_cumulative_revenue=cumulative_mrt_revenue,
                conventional_achieved_capacity_per_day=conv_plan.achieved_capacity_per_day,
                mrt_achieved_capacity_per_day=mrt_plan_core.achieved_capacity_per_day,
                conventional_retained_activity_pct=conv_plan.retained_activity_pct,
                mrt_retained_activity_pct=mrt_plan_core.retained_activity_pct,
                conventional_production_expansion_pct=conv_plan.required_production_increase_pct,
                mrt_production_expansion_pct=mrt_plan_core.required_production_increase_pct,
                capex_difference=cumulative_capex_difference,
                cumulative_economic_difference=cumulative_economic_difference,
                backbone_charged_this_phase=backbone_charged_this_phase,
                cumulative_departments_connected=len(connected_services),
                cumulative_endpoints=cumulative_endpoints,
                cumulative_guideway_segments=cumulative_guideway_segments,
                cumulative_connected_rooms=cumulative_connected_rooms,
                cumulative_supported_patients_per_day=cumulative_target_patients,
                network_utilization_pct=network_utilization_pct,
            )
        )

        network_state_rows.append(
            {
                "phase": phase.phase_name,
                "year": phase.year,
                "backbone_installed": backbone_installed,
                "cumulative_guideway_segments": cumulative_guideway_segments,
                "cumulative_endpoints": cumulative_endpoints,
                "cumulative_connected_rooms": cumulative_connected_rooms,
                "cumulative_vertical_transitions": cumulative_vertical_transitions,
                "cumulative_building_connections": cumulative_building_connections,
                "cumulative_rooms_renovated": sum(
                    float(r["incremental_quantity"]) for r in phase_ledger
                    if r["pathway"] == "MRT" and r["component"] == "Existing room renovation or modification"
                ),
                "cumulative_new_rooms_constructed": sum(
                    float(r["incremental_quantity"]) for r in phase_ledger
                    if r["pathway"] == "MRT" and r["component"] == "New room construction"
                ),
                "cumulative_conventional_scanners": int(conventional_state["scanners"]),
                "cumulative_mrt_scanners": int(mrt_state["scanners"]),
                "cumulative_conventional_usable_doses_per_day": float(conventional_state["usable_doses"]),
                "cumulative_mrt_usable_doses_per_day": float(mrt_state["usable_doses"]),
                "conventional_cyclotron_installed": bool(conventional_state["has_cyclotron"]),
                "mrt_cyclotron_installed": bool(mrt_state["has_cyclotron"]),
            }
        )

    capex_crossover_phase: str | None = None
    capex_crossover_year: int | None = None
    economic_crossover_phase: str | None = None
    economic_crossover_year: int | None = None

    for row in phase_results:
        if capex_crossover_phase is None and row.mrt_cumulative_capex <= row.conventional_cumulative_capex:
            capex_crossover_phase = row.phase_name
            capex_crossover_year = row.year
        if economic_crossover_phase is None and row.cumulative_economic_difference >= 0:
            economic_crossover_phase = row.phase_name
            economic_crossover_year = row.year

    if capex_crossover_phase is None:
        capex_crossover_summary = "No CapEx crossover within the modeled study horizon."
    else:
        capex_crossover_summary = f"CapEx crossover occurs in {capex_crossover_phase} (Year {capex_crossover_year})."

    allocated_backbone = assumptions.shared_backbone_cost / max(1, len(connected_services))

    return SharedNetworkReport(
        network_profile=profile,
        development_phases=phases_sorted,
        assumptions=assumptions,
        phase_results=phase_results,
        phase_ledger=phase_ledger,
        network_state=network_state_rows,
        cumulative_conventional_capex=cumulative_conventional_capex,
        cumulative_mrt_capex=cumulative_mrt_capex,
        cumulative_conventional_revenue=cumulative_conventional_revenue,
        cumulative_mrt_revenue=cumulative_mrt_revenue,
        capex_crossover_phase=capex_crossover_phase,
        capex_crossover_year=capex_crossover_year,
        capex_crossover_summary=capex_crossover_summary,
        economic_crossover_phase=economic_crossover_phase,
        economic_crossover_year=economic_crossover_year,
        allocated_backbone_cost_per_service_group=allocated_backbone,
    )


def shared_report_summary_dataframe(report: SharedNetworkReport) -> pd.DataFrame:
    horizon_years = report.assumptions.analysis_years
    return pd.DataFrame(
        [
            {"Metric": "Study horizon", "Value": f"{horizon_years} years"},
            {"Metric": "Number of phases", "Value": len(report.development_phases)},
            {"Metric": "Service groups connected", "Value": len({p.service_group for p in report.development_phases})},
            {"Metric": "Conventional cumulative CapEx", "Value": report.cumulative_conventional_capex},
            {"Metric": "MRT cumulative CapEx", "Value": report.cumulative_mrt_capex},
            {"Metric": "CapEx difference (MRT - Conventional)", "Value": report.cumulative_mrt_capex - report.cumulative_conventional_capex},
            {"Metric": "CapEx crossover", "Value": report.capex_crossover_summary},
            {"Metric": "Conventional cumulative revenue", "Value": report.cumulative_conventional_revenue},
            {"Metric": "MRT cumulative revenue", "Value": report.cumulative_mrt_revenue},
            {"Metric": "Allocated view - backbone cost per connected service group", "Value": report.allocated_backbone_cost_per_service_group},
            {"Metric": "Allocated view note", "Value": "Allocated view - not an additional charge."},
        ]
    )


def shared_report_phase_dataframe(report: SharedNetworkReport) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for phase in report.phase_results:
        rows.append(
            {
                "Phase": phase.phase_name,
                "Year": phase.year,
                "Service Group": phase.service_group,
                "Conventional Incremental CapEx": phase.conventional_incremental_capex,
                "MRT Incremental CapEx": phase.mrt_incremental_capex,
                "Phase CapEx Difference": phase.phase_capex_difference,
                "Conventional Cumulative CapEx": phase.conventional_cumulative_capex,
                "MRT Cumulative CapEx": phase.mrt_cumulative_capex,
                "Cumulative CapEx Difference": phase.cumulative_capex_difference,
                "Conventional Revenue": phase.conventional_phase_revenue,
                "MRT Revenue": phase.mrt_phase_revenue,
                "Conventional Cumulative Revenue": phase.conventional_cumulative_revenue,
                "MRT Cumulative Revenue": phase.mrt_cumulative_revenue,
                "CapEx Difference": phase.capex_difference,
                "Cumulative Economic Difference": phase.cumulative_economic_difference,
                "Conventional Achieved Capacity/day": phase.conventional_achieved_capacity_per_day,
                "MRT Achieved Capacity/day": phase.mrt_achieved_capacity_per_day,
                "Conventional Retained Activity %": phase.conventional_retained_activity_pct,
                "MRT Retained Activity %": phase.mrt_retained_activity_pct,
                "Conventional Production Expansion %": phase.conventional_production_expansion_pct,
                "MRT Production Expansion %": phase.mrt_production_expansion_pct,
                "Backbone Charged This Phase": "Yes" if phase.backbone_charged_this_phase else "No",
                "Cumulative Endpoints": phase.cumulative_endpoints,
                "Cumulative Guideway Segments": phase.cumulative_guideway_segments,
                "Cumulative Connected Rooms": phase.cumulative_connected_rooms,
                "Cumulative Supported Patients/day": phase.cumulative_supported_patients_per_day,
                "Network Utilization %": phase.network_utilization_pct,
            }
        )
    return pd.DataFrame(rows)


def shared_report_assumptions_dataframe(report: SharedNetworkReport) -> pd.DataFrame:
    defaults = SharedNetworkAssumptions()
    rows: list[dict[str, object]] = []
    for key in NETWORK_ASSUMPTION_FIELDS:
        applied = getattr(report.assumptions, key)
        baseline = getattr(defaults, key)
        rows.append(
            {
                "Assumption": NETWORK_ASSUMPTION_LABELS[key],
                "Applied Value": applied,
                "Standard Value": baseline,
                "Status": "User adjusted" if applied != baseline else "Standard default",
            }
        )
    return pd.DataFrame(rows)


def shared_report_ledger_dataframe(report: SharedNetworkReport) -> pd.DataFrame:
    return pd.DataFrame(report.phase_ledger)


def shared_report_network_state_dataframe(report: SharedNetworkReport) -> pd.DataFrame:
    return pd.DataFrame(report.network_state)


def shared_report_csv_dataframe(report: SharedNetworkReport) -> pd.DataFrame:
    return shared_report_phase_dataframe(report)


def _apply_currency_format(workbook, sheet_name: str, headers: set[str]) -> None:
    ws = workbook[sheet_name]
    header_to_col = {
        ws.cell(row=1, column=col_idx).value: col_idx
        for col_idx in range(1, ws.max_column + 1)
    }
    for header in headers:
        col_idx = header_to_col.get(header)
        if not col_idx:
            continue
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0"
                cell.alignment = Alignment(horizontal="right")


def shared_report_excel_bytes(report: SharedNetworkReport) -> bytes:
    summary_df = shared_report_summary_dataframe(report)
    phase_df = shared_report_phase_dataframe(report)
    assumptions_df = shared_report_assumptions_dataframe(report)
    ledger_df = shared_report_ledger_dataframe(report)
    network_state_df = shared_report_network_state_dataframe(report)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        phase_df.to_excel(writer, index=False, sheet_name="Phase Comparison")
        assumptions_df.to_excel(writer, index=False, sheet_name="Assumptions")
        ledger_df.to_excel(writer, index=False, sheet_name="Phase Ledger")
        network_state_df.to_excel(writer, index=False, sheet_name="Network State")

        workbook = writer.book
        _apply_currency_format(
            workbook,
            "Summary",
            {
                "Value",
            },
        )
        _apply_currency_format(
            workbook,
            "Phase Comparison",
            {
                "Conventional Incremental CapEx",
                "MRT Incremental CapEx",
                "Conventional Cumulative CapEx",
                "MRT Cumulative CapEx",
                "Conventional Revenue",
                "MRT Revenue",
                "Conventional Cumulative Revenue",
                "MRT Cumulative Revenue",
                "CapEx Difference",
                "Cumulative Economic Difference",
            },
        )
        _apply_currency_format(
            workbook,
            "Assumptions",
            {
                "Applied Value",
                "Standard Value",
            },
        )
        _apply_currency_format(
            workbook,
            "Phase Ledger",
            {
                "unit_cost",
                "phase_subtotal",
                "cumulative_subtotal",
            },
        )

    return buffer.getvalue()
