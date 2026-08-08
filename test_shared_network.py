import io
import math

from openpyxl import load_workbook
import pandas as pd

from diagnostics import load_radionuclide_half_lives
from models import DevelopmentPhase, NetworkProfile, PlannerAssumptions, PlannerInputs, SharedNetworkAssumptions
from optimization import conventional
from shared_network import (
    default_development_phases,
    default_network_profile,
    run_shared_network_study,
    shared_report_csv_dataframe,
    shared_report_excel_bytes,
    shared_report_phase_dataframe,
)


def _reference_profile() -> NetworkProfile:
    return default_network_profile()


def _reference_phases() -> list[DevelopmentPhase]:
    return default_development_phases()


def _lookup() -> dict[str, float]:
    return load_radionuclide_half_lives()


def test_part1_baseline_is_unchanged_by_shared_network_module():
    assumptions = PlannerAssumptions()
    inputs = PlannerInputs(
        project_name="Part1 Baseline",
        current_patients_per_day=100.0,
        target_patients_per_day=180.0,
        maximum_expected_demand_per_day=180.0,
        current_scanners=3,
        current_injection_rooms=6,
        current_uptake_rooms=6,
        has_existing_cyclotron=True,
        current_usable_doses_per_day=120.0,
        current_average_transport_min=20.0,
        mrt_transport_min=0.5,
        existing_mrt_connectable_rooms=2,
        representative_radionuclide="F-18",
        representative_half_life_min=None,
    )
    plan = conventional(inputs, assumptions, 109.8)

    assert math.isclose(plan.achieved_capacity_per_day, 183.6)
    assert math.isclose(plan.retained_activity_pct, 88.13889028316868)
    assert math.isclose(plan.required_production_increase_pct, 70.18594120947827)
    assert math.isclose(plan.capex, 14_250_000.0)


def test_shared_network_is_executable_and_backbone_is_charged_once():
    report = run_shared_network_study(
        _reference_profile(),
        _reference_phases(),
        SharedNetworkAssumptions(),
        _lookup(),
    )

    assert len(report.phase_results) >= 3
    backbone_count = sum(1 for row in report.phase_results if row.backbone_charged_this_phase)
    assert backbone_count == 1
    assert report.phase_results[0].backbone_charged_this_phase is True
    assert all(not row.backbone_charged_this_phase for row in report.phase_results[1:])


def test_state_carries_forward_and_ledger_reconciles():
    report = run_shared_network_study(
        _reference_profile(),
        _reference_phases(),
        SharedNetworkAssumptions(),
        _lookup(),
    )

    phase_df = shared_report_phase_dataframe(report)
    ledger = pd.DataFrame(report.phase_ledger)
    network = pd.DataFrame(report.network_state)

    assert math.isclose(
        phase_df["Conventional Incremental CapEx"].sum(),
        report.cumulative_conventional_capex,
    )
    assert math.isclose(
        phase_df["MRT Incremental CapEx"].sum(),
        report.cumulative_mrt_capex,
    )

    mrt_guideway = ledger[(ledger["pathway"] == "MRT") & (ledger["component"] == "Guideway segments")]
    mrt_endpoints = ledger[(ledger["pathway"] == "MRT") & (ledger["component"] == "Endpoints")]
    assert all(mrt_guideway["incremental_quantity"] >= 0)
    assert all(mrt_endpoints["incremental_quantity"] >= 0)
    assert all(
        math.isclose(
            float(row.required_total_quantity_or_state) - float(row.opening_quantity_or_state),
            float(row.incremental_quantity),
        )
        for row in mrt_guideway.itertuples()
    )
    assert all(
        math.isclose(
            float(row.required_total_quantity_or_state) - float(row.opening_quantity_or_state),
            float(row.incremental_quantity),
        )
        for row in mrt_endpoints.itertuples()
    )

    assert list(network["cumulative_mrt_scanners"]) == sorted(list(network["cumulative_mrt_scanners"]))
    assert list(network["cumulative_mrt_usable_doses_per_day"]) == sorted(list(network["cumulative_mrt_usable_doses_per_day"]))

    for phase_name in phase_df["Phase"]:
        phase_rows = ledger[ledger["phase"] == phase_name]
        conv_phase_ledger = phase_rows[phase_rows["pathway"] == "Conventional"]["phase_subtotal"].sum()
        mrt_phase_ledger = phase_rows[phase_rows["pathway"] == "MRT"]["phase_subtotal"].sum()
        row = phase_df[phase_df["Phase"] == phase_name].iloc[0]
        assert math.isclose(float(conv_phase_ledger), float(row["Conventional Incremental CapEx"]))
        assert math.isclose(float(mrt_phase_ledger), float(row["MRT Incremental CapEx"]))


def test_room_connection_and_construction_are_not_forced_double_charge():
    profile = _reference_profile()
    phases = [
        DevelopmentPhase(
            phase_name="Phase 1",
            year=0,
            service_group="PET",
            representative_radionuclide="F-18",
            incremental_target_patients_per_day=10.0,
            maximum_expected_demand_per_day=10.0,
            existing_rooms_to_connect=6,
            new_rooms_to_construct=0,
            cumulative_guideway_segments_required=1,
            cumulative_endpoints_required=4,
            cumulative_vertical_transitions_required=0,
            cumulative_building_connections_required=0,
            new_rooms_require_connection_modification=False,
        ),
        DevelopmentPhase(
            phase_name="Phase 2",
            year=1,
            service_group="Theranostics",
            representative_radionuclide="Ga-68",
            incremental_target_patients_per_day=10.0,
            maximum_expected_demand_per_day=10.0,
            existing_rooms_to_connect=2,
            new_rooms_to_construct=0,
            cumulative_guideway_segments_required=2,
            cumulative_endpoints_required=6,
        ),
        DevelopmentPhase(
            phase_name="Phase 3",
            year=2,
            service_group="SPECT",
            representative_radionuclide="Tc-99m",
            incremental_target_patients_per_day=10.0,
            maximum_expected_demand_per_day=10.0,
            existing_rooms_to_connect=1,
            new_rooms_to_construct=0,
            cumulative_guideway_segments_required=3,
            cumulative_endpoints_required=8,
        ),
    ]

    report = run_shared_network_study(profile, phases, SharedNetworkAssumptions(), _lookup())
    ledger = pd.DataFrame(report.phase_ledger)

    phase1_construct = ledger[
        (ledger["phase"] == "Phase 1")
        & (ledger["pathway"] == "MRT")
        & (ledger["component"] == "New room construction")
    ]["incremental_quantity"].iloc[0]
    phase1_connect = ledger[
        (ledger["phase"] == "Phase 1")
        & (ledger["pathway"] == "MRT")
        & (ledger["component"] == "Room connection or modification")
    ]["incremental_quantity"].iloc[0]
    phase1_renovate = ledger[
        (ledger["phase"] == "Phase 1")
        & (ledger["pathway"] == "MRT")
        & (ledger["component"] == "Existing room renovation or modification")
    ]["incremental_quantity"].iloc[0]

    assert phase1_construct == 0
    assert phase1_connect == 6
    assert phase1_renovate == 0


def test_room_cost_classification_new_vs_renovation_vs_connection():
    profile = _reference_profile()
    phases = [
        DevelopmentPhase(
            phase_name="Phase 1",
            year=0,
            service_group="PET",
            representative_radionuclide="F-18",
            incremental_target_patients_per_day=20.0,
            maximum_expected_demand_per_day=20.0,
            existing_rooms_to_connect=2,
            existing_rooms_to_renovate=3,
            new_rooms_to_construct=1,
            mrt_new_rooms_requiring_connection_modification=0,
            cumulative_guideway_segments_required=2,
            cumulative_endpoints_required=6,
        ),
        DevelopmentPhase(
            phase_name="Phase 2",
            year=1,
            service_group="Theranostics",
            representative_radionuclide="Ga-68",
            incremental_target_patients_per_day=15.0,
            maximum_expected_demand_per_day=15.0,
            existing_rooms_to_connect=1,
            existing_rooms_to_renovate=1,
            new_rooms_to_construct=1,
            cumulative_guideway_segments_required=3,
            cumulative_endpoints_required=8,
        ),
        DevelopmentPhase(
            phase_name="Phase 3",
            year=2,
            service_group="SPECT",
            representative_radionuclide="Tc-99m",
            incremental_target_patients_per_day=10.0,
            maximum_expected_demand_per_day=10.0,
            existing_rooms_to_connect=1,
            existing_rooms_to_renovate=0,
            new_rooms_to_construct=0,
            cumulative_guideway_segments_required=3,
            cumulative_endpoints_required=9,
        ),
    ]
    assumptions = SharedNetworkAssumptions(
        new_room_construction_capex=100_000.0,
        room_renovation_modification_capex=20_000.0,
        room_connection_modification_capex=5_000.0,
    )

    report = run_shared_network_study(profile, phases, assumptions, _lookup())
    ledger = pd.DataFrame(report.phase_ledger)

    p1_new = ledger[(ledger["phase"] == "Phase 1") & (ledger["pathway"] == "MRT") & (ledger["component"] == "New room construction")]["incremental_quantity"].iloc[0]
    p1_ren = ledger[(ledger["phase"] == "Phase 1") & (ledger["pathway"] == "MRT") & (ledger["component"] == "Existing room renovation or modification")]["incremental_quantity"].iloc[0]
    p1_conn = ledger[(ledger["phase"] == "Phase 1") & (ledger["pathway"] == "MRT") & (ledger["component"] == "Room connection or modification")]["incremental_quantity"].iloc[0]

    assert p1_new == 1
    assert p1_ren == 3
    assert p1_conn == 2


def test_infrastructure_carry_forward_and_incremental_phase_spend_only():
    report = run_shared_network_study(
        _reference_profile(),
        _reference_phases(),
        SharedNetworkAssumptions(),
        _lookup(),
    )
    ledger = pd.DataFrame(report.phase_ledger)

    for component in [
        "Guideway segments",
        "Endpoints",
        "Vertical transitions",
        "Building or floor connections",
    ]:
        rows = ledger[(ledger["pathway"] == "MRT") & (ledger["component"] == component)].sort_values("year")
        prev_required = None
        for row in rows.itertuples():
            if prev_required is not None:
                assert math.isclose(float(row.opening_quantity_or_state), float(prev_required))
            assert math.isclose(
                float(row.required_total_quantity_or_state) - float(row.opening_quantity_or_state),
                float(row.incremental_quantity),
            )
            prev_required = float(row.required_total_quantity_or_state)


def test_no_room_double_counting_within_phase_for_mrt():
    report = run_shared_network_study(
        _reference_profile(),
        _reference_phases(),
        SharedNetworkAssumptions(),
        _lookup(),
    )
    ledger = pd.DataFrame(report.phase_ledger)

    for phase_name in ledger["phase"].unique():
        phase_rows = ledger[(ledger["phase"] == phase_name) & (ledger["pathway"] == "MRT")]
        new_qty = float(phase_rows[phase_rows["component"] == "New room construction"]["incremental_quantity"].iloc[0])
        ren_qty = float(phase_rows[phase_rows["component"] == "Existing room renovation or modification"]["incremental_quantity"].iloc[0])
        conn_qty = float(phase_rows[phase_rows["component"] == "Room connection or modification"]["incremental_quantity"].iloc[0])

        conn_row = phase_rows[phase_rows["component"] == "Room connection or modification"].iloc[0]
        new_conn_qty = float(conn_row.get("new_rooms_connection_modified", 0.0))
        existing_conn_qty = float(conn_row.get("existing_rooms_connected", 0.0))

        assert new_conn_qty <= new_qty
        assert existing_conn_qty <= conn_qty
        assert conn_qty >= new_conn_qty + existing_conn_qty
        assert ren_qty >= 0.0


def test_decay_affects_phase_production_requirements_and_crossover_paths_work():
    profile = _reference_profile()
    phases = _reference_phases()

    fast = SharedNetworkAssumptions(mrt_transport_min=0.5)
    slow = SharedNetworkAssumptions(mrt_transport_min=20.0)

    report_fast = run_shared_network_study(profile, phases, fast, _lookup())
    report_slow = run_shared_network_study(profile, phases, slow, _lookup())

    assert report_fast.phase_results[0].mrt_production_expansion_pct <= report_slow.phase_results[0].mrt_production_expansion_pct

    force_crossover = SharedNetworkAssumptions(
        shared_backbone_cost=0.0,
        guideway_segment_capex=0.0,
        endpoint_capex=0.0,
        vertical_transition_capex=0.0,
        building_connection_capex=0.0,
        room_connection_modification_capex=0.0,
    )
    no_crossover = SharedNetworkAssumptions(
        shared_backbone_cost=100_000_000.0,
        guideway_segment_capex=2_000_000.0,
        endpoint_capex=50_000.0,
        vertical_transition_capex=1_000_000.0,
        building_connection_capex=1_000_000.0,
    )

    report_cross = run_shared_network_study(profile, phases, force_crossover, _lookup())
    report_no_cross = run_shared_network_study(profile, phases, no_crossover, _lookup())

    assert report_cross.capex_crossover_year is not None
    assert report_no_cross.capex_crossover_year is None
    assert "No CapEx crossover" in report_no_cross.capex_crossover_summary


def test_csv_excel_traceability_currency_format_and_no_stale_results_on_rerun():
    profile = _reference_profile()
    phases = _reference_phases()
    assumptions = SharedNetworkAssumptions()

    report = run_shared_network_study(profile, phases, assumptions, _lookup())
    phase_df = shared_report_phase_dataframe(report)
    csv_df = shared_report_csv_dataframe(report)
    book = shared_report_excel_bytes(report)

    assert list(csv_df["Conventional Production Expansion %"]) == list(phase_df["Conventional Production Expansion %"])
    assert list(csv_df["MRT Production Expansion %"]) == list(phase_df["MRT Production Expansion %"])

    excel_phase = pd.read_excel(io.BytesIO(book), sheet_name="Phase Comparison")
    assert all(
        math.isclose(float(a), float(b))
        for a, b in zip(
            list(excel_phase["Conventional Production Expansion %"]),
            list(phase_df["Conventional Production Expansion %"]),
        )
    )
    assert all(
        math.isclose(float(a), float(b))
        for a, b in zip(
            list(excel_phase["MRT Production Expansion %"]),
            list(phase_df["MRT Production Expansion %"]),
        )
    )

    wb = load_workbook(io.BytesIO(book))
    ws = wb["Phase Comparison"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    capex_col = headers.index("Conventional Incremental CapEx") + 1
    assert ws.cell(row=2, column=capex_col).number_format == "$#,##0"

    phases_edited = list(phases)
    last = phases_edited[-1]
    phases_edited[-1] = DevelopmentPhase(
        phase_name=last.phase_name,
        year=last.year,
        service_group=last.service_group,
        representative_radionuclide=last.representative_radionuclide,
        incremental_target_patients_per_day=last.incremental_target_patients_per_day + 20.0,
        maximum_expected_demand_per_day=last.maximum_expected_demand_per_day + 20.0,
        existing_rooms_to_connect=last.existing_rooms_to_connect,
        new_rooms_to_construct=last.new_rooms_to_construct,
        cumulative_guideway_segments_required=last.cumulative_guideway_segments_required,
        cumulative_endpoints_required=last.cumulative_endpoints_required,
        cumulative_vertical_transitions_required=last.cumulative_vertical_transitions_required,
        cumulative_building_connections_required=last.cumulative_building_connections_required,
        additional_scanners_manual=last.additional_scanners_manual,
        production_demand_multiplier=last.production_demand_multiplier,
        conventional_new_cyclotron_required=last.conventional_new_cyclotron_required,
        mrt_new_cyclotron_required=last.mrt_new_cyclotron_required,
        can_use_existing_backbone_capacity=last.can_use_existing_backbone_capacity,
        new_rooms_require_connection_modification=last.new_rooms_require_connection_modification,
    )

    report_edited = run_shared_network_study(profile, phases_edited, assumptions, _lookup())
    assert not math.isclose(report_edited.cumulative_conventional_capex, report.cumulative_conventional_capex)
    assert report_edited.phase_results[-1].cumulative_supported_patients_per_day > report.phase_results[-1].cumulative_supported_patients_per_day
